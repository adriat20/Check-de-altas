# -*- coding: utf-8 -*-

from __future__ import annotations
import re, sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import unicodedata
from pandas._libs.tslibs.timestamps import Timestamp

# ---------------- CONFIG ----------------------------------------------------
BASE_DIR = Path(r"C:\Users\ofici\OneDrive\ESCRITORIO IBERDROLA\PROGRAMACION\Proyecto_Check_Altas")
#BASE_DIR = Path(r"C:\Users\X\OneDrive\ESCRITORIO IBERDROLA\PROGRAMACION\Proyecto_Check_Altas")
SRC_XLS  = BASE_DIR / "2025_TRAMITACION_DE_ALTAS.xlsx"


if len(sys.argv) >= 4:
    SHEETS = [sys.argv[3].strip().upper()]
else:
    hoja = input("📄 ¿Qué mes quieres analizar?: ").strip().upper()
    SHEETS = [hoja]

PLANES    = ["2,0 TD_1", "2,0 TD_2", "2,0 TD_3", "3,0 TD"]
SERVS     = {"PIH":["PIH"], "PEH+":["PEH+"], "UUEEn/UUEE":["UUEEN","UUEE"], "PTG":["PTG"]}
OFERTA    = "EXCLUSIVO 10% TF/TV"

# -------------- HELPERS -----------------------------------------------------
def contains(series: pd.Series, toks: list[str]):
    pat = r"(?<![A-Z0-9])(?:" + "|".join(map(re.escape, toks)) + r")(?![A-Z0-9])"
    return series.str.contains(pat, na=False, regex=True)

def auto_width(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            max(len(str(c.value)) for c in col if c.value) + 2
        )

def ask_date(msg: str):
    while True:
        txt = input(f"{msg} (dd-mm-aaaa): ").strip()
        try:
            return pd.to_datetime(txt, format="%d-%m-%Y")
        except ValueError:
            print("❌ Formato incorrecto.")

# -------------- LOAD --------------------------------------------------------
print("⏳ Cargando hoja(s):", ", ".join(SHEETS))
raw = pd.concat(
    [pd.read_excel(SRC_XLS, sheet_name=s, engine="openpyxl") for s in SHEETS],
    ignore_index=True,
).drop_duplicates()

# ─── Normaliza cabeceras (tildes, espacios, mayúsculas) ─────────────────────
def sin_tildes(txt):
    return "".join(c for c in unicodedata.normalize("NFKD", txt)
                   if not unicodedata.combining(c))

raw.columns = [sin_tildes(col).upper().strip() for col in raw.columns]

# ─── Renombra columnas erróneas (por si aparece mal escrito) ───────────────
raw.rename(columns={
    "CODIGO COMERCIAL": "CODIGO COMERCIAL",
    "CODICO COMERCIAL": "CODIGO COMERCIAL"
}, inplace=True)

# ─── FILTRA FILAS CON CABECERAS PEGADAS O VACÍAS ───────────────────────────
header_like = set(raw.columns)
header_like.update({"DOC. SUBIDA"})
raw = raw[~raw["COLABORADOR"].str.upper().isin(header_like)]
raw = raw[raw["COLABORADOR"].str.strip() != ""]

# ─── Normaliza campos de texto clave ───────────────────────────────────────
for c in ["PUNTO ATENCION","SERVICIOS","COMUNIDAD","OFERTA PRESENTADA","COLABORADOR"]:
    raw[c] = raw[c].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)

    # ─── Convierte a fecha y elimina valores no escalares en CAIDAS ────────────
for c in ["FECHA FIRMA", "FECHA ALTA", "CAIDAS"]:
    raw[c] = pd.to_datetime(raw[c], errors="coerce")

from pandas._libs.tslibs.timestamps import Timestamp
raw.loc[~raw["CAIDAS"].apply(lambda x: isinstance(x, Timestamp)), "CAIDAS"] = pd.NaT
raw.loc[~raw["FECHA ALTA"].apply(lambda x: isinstance(x, Timestamp)), "FECHA ALTA"] = pd.NaT

# -------------- DATES -------------------------------------------------------
if len(sys.argv) >= 3:
    d_ini = pd.to_datetime(sys.argv[1], format="%d-%m-%Y")
    d_fin = pd.to_datetime(sys.argv[2], format="%d-%m-%Y")
else:
    print("⚠️ Sin fechas → pedir.")
    d_ini = ask_date("Desde")
    d_fin = ask_date("Hasta")

mask_firma = raw["FECHA FIRMA"].between(d_ini, d_fin, "both")
mask_caida = raw["CAIDAS"].between(d_ini, d_fin, "both")
mask_alta_null  = raw["FECHA ALTA"].isna()
mask_caida_null = raw["CAIDAS"].isna()

ALTAS = raw[mask_firma & raw["CAIDAS"].isna()]
# Recupera texto original de FECHA ALTA
raw["FECHA ALTA ORIGINAL"] = (
    pd.read_excel(SRC_XLS, sheet_name=hoja, engine="openpyxl")
      ["FECHA ALTA"]
)
INCID = raw[
    mask_firma &
    raw["CAIDAS"].isna() &
    raw["FECHA ALTA"].isna() &             # No se pudo convertir a fecha
    raw["FECHA ALTA ORIGINAL"].notna()     # Pero hay texto
]

BAJAS = raw[mask_caida]

# -------------- TOTAL_GLOBAL ------------------------------------------------
rows = []
raw["CODIGO COMERCIAL"] = raw["CODIGO COMERCIAL"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)


# Códigos comerciales por ubicación

#MIERES
M_CODE = ["YB19010-ANA-3188168", "YB99670-ADRIAN-155292"]
#LENA
L_CODE = ["YB33990-ELI-3189791"]
#PYMES
P_CODE = ["YA8541- GERAR- 3184474"]

def is_mieres(df):
    return df["CODIGO COMERCIAL"].isin(M_CODE)

def is_lena(df):
    return df["CODIGO COMERCIAL"].isin(L_CODE)

def is_pymes(df):
    return df["CODIGO COMERCIAL"].isin(P_CODE) & df["PLAN"].isin(["2,0 TD_3", "3,0 TD"])

def add(tipo, df_a, df_b):
    rows.append({
        "TIPO": tipo,
        "ALTAS": df_a.shape[0],
        "BAJAS": df_b.shape[0],
        "NO_ASTURIAS": df_a[df_a["COMUNIDAD"]!="ASTURIAS"].shape[0],
        "TOTALES": df_a.shape[0] - df_b.shape[0] - df_a[df_a["COMUNIDAD"]!="ASTURIAS"].shape[0],
        "ALTAS_LENA": df_a[is_lena(df_a)].shape[0],
        "BAJAS_LENA": df_b[is_lena(df_b)].shape[0],
        "ALTAS_MIERES": df_a[is_mieres(df_a)].shape[0],
        "BAJAS_MIERES": df_b[is_mieres(df_b)].shape[0],
        "ALTAS_PYMES": df_a[is_pymes(df_a)].shape[0],
        "BAJAS_PYMES": df_b[is_pymes(df_b)].shape[0],

    })


for p in PLANES:
    add(p,
        ALTAS[ALTAS["PLAN"].str.startswith(p,na=False)],
        BAJAS[BAJAS["PLAN"].str.startswith(p,na=False)]
    )
add("Plan Exclusivo 10%",
    ALTAS[contains(ALTAS["OFERTA PRESENTADA"], [OFERTA])],
    BAJAS[contains(BAJAS["OFERTA PRESENTADA"], [OFERTA])]
)
for k,toks in SERVS.items():
    add(k,
        ALTAS[contains(ALTAS["SERVICIOS"], toks)],
        BAJAS[contains(BAJAS["SERVICIOS"], toks)]
    )

add("ALTAS CON INCIDENCIA", INCID, INCID)



total_global = pd.DataFrame(rows)
print(total_global[["TIPO","ALTAS","BAJAS"]])

# -------------- POR_COLAB ---------------------------------------------------
plan_alt = (
    ALTAS[ALTAS["PLAN"].isin(PLANES)]
    .groupby(["COLABORADOR","PLAN"]).size()
    .unstack(fill_value=0).reindex(columns=PLANES, fill_value=0)
)
plan_baj = (
    BAJAS[BAJAS["PLAN"].isin(PLANES)]
    .groupby(["COLABORADOR","PLAN"]).size()
    .unstack(fill_value=0).reindex(columns=PLANES, fill_value=0)
)
plan_alt.columns = [f"PLAN_{c}_ALTA"  for c in plan_alt.columns]
plan_baj.columns = [f"PLAN_{c}_CAIDA" for c in plan_baj.columns]

serv_alt = ALTAS.groupby("COLABORADOR").apply(
    lambda df: pd.Series({
        f"SERVICIO_{k}_ALTA": contains(df["SERVICIOS"], toks).sum()
        for k,toks in SERVS.items()
    })
)
serv_baj = BAJAS.groupby("COLABORADOR").apply(
    lambda df: pd.Series({
        f"SERVICIO_{k}_CAIDA": contains(df["SERVICIOS"], toks).sum()
        for k,toks in SERVS.items()
    })
)

of_alt = (ALTAS[contains(ALTAS["OFERTA PRESENTADA"],[OFERTA])]
          .groupby("COLABORADOR").size().to_frame(f"OFERTA_{OFERTA}_ALTA"))
of_baj = (BAJAS[contains(BAJAS["OFERTA PRESENTADA"],[OFERTA])]
          .groupby("COLABORADOR").size().to_frame(f"OFERTA_{OFERTA}_CAIDA"))

por_colab = (plan_alt.join(plan_baj,how="outer")
                     .join(of_alt,how="outer")
                     .join(of_baj,how="outer")
                     .join(serv_alt,how="outer")
                     .join(serv_baj,how="outer")
                     .fillna(0).astype(int))
por_colab_t = por_colab.T.reset_index()
por_colab_t.columns = ["INDICADOR"] + por_colab_t.columns[1:].tolist()
# ── FILTRA solo indicadores válidos ────────────────────────────────────────
valid = por_colab_t["INDICADOR"].str.match(r"^(PLAN_|OFERTA_|SERVICIO_)")
por_colab_t = por_colab_t[valid].reset_index(drop=True)
# -------------- EXPORT ------------------------------------------------------
out = BASE_DIR / f"Resumen_colaboradores_{datetime.today():%Y-%m-%d}.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    por_colab_t.to_excel(writer, sheet_name="POR_COLABORADOR", index=False)
    total_global.to_excel(writer, sheet_name="TOTAL_GLOBAL",   index=False)
print(f"💾 {out}")

# -------------- FORMAT ------------------------------------------------------
wb = load_workbook(out)

# Colores y estilos
fills = {
    "head":  PatternFill("solid", fgColor="B7E1CD"),  # verde cabecera
    "alta":  PatternFill("solid", fgColor="C6EFCE"),  # 🟢 verde claro
    "alta_loc": PatternFill("solid", fgColor="C6EFCE"),  # 🟢 verde claro (altas Mieres/Lena/Pymes)
    "baja":  PatternFill("solid", fgColor="FFC7CE"),  # 🔴 rojo claro
    "inci":  PatternFill("solid", fgColor="FFF599"),  # 🟡 amarillo
    "title": PatternFill("solid", fgColor="FBE4D5"),
    "total": PatternFill("solid", fgColor="BDD7EE"),
    "leyenda": PatternFill("solid", fgColor="FFF599"),  # Amarillo específico para leyenda
}
fonts = {"head":Font(bold=True), "title":Font(bold=True, size=12)}
align = Alignment(horizontal="center", vertical="center")
border= Border(*(Side("thin") for _ in range(4)))
per_txt = f"📅 PERÍODO: {d_ini:%d-%m-%Y} → {d_fin:%d-%m-%Y}"

for sh in ["POR_COLABORADOR","TOTAL_GLOBAL"]:
    ws = wb[sh]
    
    # Inserta y formatea el título
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    hdr = ws.cell(1, 1, per_txt)
    hdr.fill, hdr.font, hdr.alignment = fills["title"], fonts["title"], align

    # Formatea encabezados
    for c in ws[2]:
        c.fill, c.font, c.alignment, c.border = fills["head"], fonts["head"], align, border

    # Formatea datos (excluyendo la leyenda)
    if sh == "TOTAL_GLOBAL":
        # Primero creamos la leyenda completamente aislada
        BLANK_ROWS = 3
        data_end_row = ws.max_row
        ws.insert_rows(data_end_row + 1, amount=BLANK_ROWS)
        
        # Crea leyenda con formato protegido
        legend_lines = [
            "LEYENDA:",
            "• 🔼 *ALTA*: Firma dentro del período y fecha de CAÍDA vacía.",
            "• 🔽 *BAJA*: Fecha de CAÍDAS dentro del período (independiente de la alta).",
            "• ⚠️ *INCIDENCIA*: Firma dentro del período sin alta válida ni caída.",
            "ℹ️ *RECUERDA*: Las altas con incidencia (RECHAZO, T/A, etc.) se muestran en amarillo y no cuentan como altas ni como bajas."
        ]
        
        # Marca filas de leyenda para excluirlas del formateo posterior
        legend_start_row = data_end_row + BLANK_ROWS + 1
        legend_rows = set(range(legend_start_row, legend_start_row + len(legend_lines)))
        
        border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
        )
        col_span = ws.max_column


        # Aplica formato a la leyenda
        for i, text in enumerate(legend_lines):
            row_num = legend_start_row + i
            ws.merge_cells(start_row=row_num, start_column=1, 
                          end_row=row_num, end_column=ws.max_column)
            cell = ws.cell(row=row_num, column=1, value=text)
            cell.fill = fills["leyenda"]
            cell.border = border_thin
            if i == 0:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        
        # Bordes gruesos para la leyenda
        thick = Side(style="medium")
        for col in range(1, ws.max_column + 1):
            top_cell = ws.cell(row=legend_start_row, column=col)
            bot_cell = ws.cell(row=legend_start_row + len(legend_lines) - 1, column=col)
            top_cell.border = Border(top=thick, left=border_thin.left,
                                    right=border_thin.right, bottom=border_thin.bottom)
            bot_cell.border = Border(bottom=thick, left=border_thin.left,
                                     right=border_thin.right, top=border_thin.top)
        
        # Ajusta el rango de formateo para excluir la leyenda
        max_data_row = data_end_row
    else:
        max_data_row = ws.max_row

    # Aplica formato solo a las filas de datos (excluyendo leyenda)
    for row_cells in ws.iter_rows(min_row=3, max_row=max_data_row):
        first = row_cells[0].value
        if first is None:
            continue
        tipo = first
        for c in row_cells:
            c.alignment, c.border = align, border
            if sh == "POR_COLABORADOR":
                c.fill = fills["alta"] if "_ALTA" in str(row_cells[0].value) else fills["baja"]
            else:
                hdr_txt = ws.cell(row=2, column=c.column).value
                if hdr_txt == "TOTALES":
                    c.fill = fills["total"]
                elif tipo == "ALTAS CON INCIDENCIA":
                    c.fill = fills["inci"]
                elif hdr_txt and hdr_txt.startswith("ALTAS_"):
                    c.fill = fills["alta_loc"]
                elif hdr_txt and hdr_txt.startswith("BAJAS_"):
                    c.fill = fills["baja"]

    auto_width(ws)
    
    if sh == "TOTAL_GLOBAL":
        max_allowed = 25
        if ws.column_dimensions['A'].width > max_allowed:
            ws.column_dimensions['A'].width = max_allowed

wb.save(out)

# -------------- HOJA TRAMITACION ------------------------------------------------------
resp = input("¿Quieres leer la hoja de TRAMITACIÓN también? (S/N): ").strip().upper()
if resp == 'S':
    print("📄 Leyendo la hoja de TRAMITACIÓN …")
    df_tram = pd.read_excel(SRC_XLS, sheet_name='TRAMITACION', engine="openpyxl")
    # — Normalizar cabeceras de df_tram —
    def sin_tildes(txt):
        return "".join(c for c in unicodedata.normalize("NFKD", txt)
                       if not unicodedata.combining(c))
    df_tram.columns = [sin_tildes(col).upper().strip() for col in df_tram.columns]

    # Carga las dos hojas siguientes tras TRAMITACION
    wb_src = load_workbook(SRC_XLS, read_only=True)
    sheet_names = wb_src.sheetnames
    idx = sheet_names.index("TRAMITACION")
    next_two = sheet_names[idx+1:idx+3]
    df_extra = pd.concat(
        [pd.read_excel(SRC_XLS, sheet_name=s, engine="openpyxl") for s in next_two],
        ignore_index=True
    )
    df_extra.columns = [sin_tildes(col).upper().strip() for col in df_extra.columns]
    for c in ["FECHA FIRMA","FECHA ALTA","CAIDAS"]:
        df_extra[c] = pd.to_datetime(df_extra[c], errors="coerce")

    # Unimos todo en df_tram
    df_tram = pd.concat([df_tram, df_extra], ignore_index=True)

    # Ahora convertimos fechas en el DataFrame resultante
    for c in ["FECHA FIRMA","FECHA ALTA","CAIDAS"]:
        df_tram[c] = pd.to_datetime(df_tram[c], errors="coerce")

    wb = load_workbook(out)
    colaboradores = df_tram['COLABORADOR'].dropna().unique()

    
    # Define los rellenos
    fill_alta = PatternFill("solid", fgColor="D5F5D3")   # verde muy claro
    fill_baja = PatternFill("solid", fgColor="FFC7CE")   # rojo muy claro
    fill_inci = PatternFill("solid", fgColor="FFF2CC")   # amarillo muy claro
    fill_firma_ant = PatternFill("solid", fgColor="C9DAF8") # azul muy claro

    legend = [
    ("🟩 ALTA",    "Firma entre d_ini y d_fin, sin caída"),
    ("🟥 BAJA",    "Caída entre d_ini y d_fin"),
    ("🟨 INCIDENCIA", "Firma en rango, sin alta válida ni caída"),
    ("🟦 SECUNDARIO", "Firma < d_ini y caída entre d_ini y hoy")]


    # Estilo de encabezado
    hdr_fill   = PatternFill("solid", fgColor="BDD7EE")
    hdr_font   = Font(bold=True, color="000000", size=12)
    hdr_align  = Alignment(horizontal="center", vertical="center")

    fecha_hoy = pd.to_datetime(datetime.today().date())
    for col in colaboradores:
        # Limpia y normaliza el nombre de hoja
        nombre = str(col).strip()[:31]
        nombre = re.sub(r"[\\/?*\\[\\]]", "_", nombre)

        cols = [
            'COLABORADOR','NOMBRE DEL CLIENTE','DNI/CIF',
            'PLAN','POTENCIA','OFERTA PRESENTADA','SERVICIOS',
            'FECHA FIRMA','FECHA ALTA','OBSERV.','CAIDAS','CHECK ALTAS',
            'VIENE GRACIAS A :','OTROS'
        ]

        col_clean = str(col).strip().upper()
        mask_col  = df_tram['COLABORADOR'].astype(str).str.strip().str.upper() == col_clean
        mask_alta = df_tram['FECHA FIRMA'].between(d_ini, d_fin, inclusive='both') & df_tram['CAIDAS'].isna()
        mask_baja = df_tram['CAIDAS'].between(d_ini, d_fin, inclusive='both')
        mask_inci = mask_alta & df_tram['FECHA ALTA'].isna()
        mask_firma_ant = ((df_tram['FECHA FIRMA'] < d_ini) & df_tram['CAIDAS'].between(d_ini, fecha_hoy, inclusive="both"))


        df_fil = df_tram[mask_col & (mask_alta | mask_baja | mask_inci | mask_firma_ant)][cols]


        ws = wb.create_sheet(title=nombre)

        # 1) Escribe encabezados y les da estilo
    for c_idx, header in enumerate(df_fil.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border  # si quieres el mismo borde que antes

        # 2) Escribe las filas y las colorea (igual que tenías)
    for r_idx, row in enumerate(df_fil.itertuples(), start=2):
        idx        = row.Index
        is_firma_ant = mask_firma_ant.loc[idx]
        is_inci    = mask_inci.loc[idx]
        is_baja    = mask_baja.loc[idx]
        for c_idx, value in enumerate(row[1:], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if is_firma_ant:
                cell.fill = fill_firma_ant
            elif is_inci:
                cell.fill = fill_inci
            elif is_baja:
                cell.fill = fill_baja
            else:
                cell.fill = fill_alta

        auto_width(ws)

    # 3) Inserta leyenda 2 filas por debajo de los datos
    start = ws.max_row + 2
    for i, (label, desc) in enumerate(legend):
        cell = ws.cell(row=start + i, column=1, value=f"{label}: {desc}")
        cell.font      = Font(italic=True)
        cell.alignment = Alignment(vertical="center")
        # si quieres color de fondo en la leyenda:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

    # Fusiona la leyenda a lo ancho
    for i in range(len(legend)):
        ws.merge_cells(start_row=start+i, start_column=1,
                       end_row=start+i,   end_column=len(cols))        

    wb.save(out)
    print(f"✅ Agregadas {len(colaboradores)} hojas de TRAMITACIÓN al archivo {out}")

print("✅ Fin.")
