# -*- coding: utf-8 -*-
"""
check_altas_resumen_fixed.py — FINAL
------------------------------------
Genera los informes POR_COLABORADOR y TOTAL_GLOBAL garantizando que las cifras
coinciden con la hoja ABRIL de 2025_TRAMITACION_DE_ALTAS.xlsx para el rango
que el usuario indique.
"""
from __future__ import annotations
import re, sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import unicodedata
from pandas._libs.tslibs.timestamps import Timestamp
import warnings
from openpyxl.utils.exceptions import InvalidFileException

# Ignorar UserWarning (incluye los de openpyxl)
warnings.filterwarnings("ignore", category=UserWarning)

# Ignorar FutureWarning
warnings.filterwarnings("ignore", category=FutureWarning)

# Ignorar DeprecationWarning
warnings.filterwarnings("ignore", category=DeprecationWarning)
# ---------------- CONFIG ----------------------------------------------------
BASE_DIR = Path(r"C:\Users\ofici\OneDrive\ESCRITORIO IBERDROLA\PROGRAMACION\Proyecto_Check_Altas")
#BASE_DIR = Path(r"C:\Users\X\OneDrive\ESCRITORIO IBERDROLA\PROGRAMACION\Proyecto_Check_Altas")
SRC_XLS  = BASE_DIR / "2025_TRAMITACION_DE_ALTAS.xlsx"


if len(sys.argv) >= 4:
    SHEETS = [sys.argv[3].strip().upper()]
else:
    hoja = input("📄 ¿Qué mes quieres analizar?: ").strip().upper()
    SHEETS = [hoja]

PLANES    = ["2,0 TD_1", "2,0 TD_2", "2,0 TD_3", "3,0 TD"]
SERVS     = {"PIH":["PIH"], "PEH+":["PEH+"], "UUEEn/UUEE":["UUEEN","UUEE"], "PTG":["PTG"]}
OFERTA    = "EXCLUSIVO 10% TF/TV"

# -------------- HELPERS -----------------------------------------------------
def contains(series: pd.Series, toks: list[str]):
    pat = r"(?<![A-Z0-9])(?:" + "|".join(map(re.escape, toks)) + r")(?![A-Z0-9])"
    return series.str.contains(pat, na=False, regex=True)

def auto_width(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            max(len(str(c.value)) for c in col if c.value) + 2
        )

def ask_date(msg: str):
    while True:
        txt = input(f"{msg} (dd-mm-aaaa): ").strip()
        try:
            return pd.to_datetime(txt, format="%d-%m-%Y")
        except ValueError:
            print("❌ Formato incorrecto.")

# -------------- LOAD --------------------------------------------------------
print("⏳ Cargando hoja(s):", ", ".join(SHEETS))
raw = pd.concat(
    [pd.read_excel(SRC_XLS, sheet_name=s, engine="openpyxl") for s in SHEETS],
    ignore_index=True,
).drop_duplicates()

# ─── Normaliza cabeceras (tildes, espacios, mayúsculas) ─────────────────────
def sin_tildes(txt):
    return "".join(c for c in unicodedata.normalize("NFKD", txt)
                   if not unicodedata.combining(c))

raw.columns = [sin_tildes(col).upper().strip() for col in raw.columns]

# ─── Renombra columnas erróneas (por si aparece mal escrito) ───────────────
raw.rename(columns={
    "CODIGO COMERCIAL": "CODIGO COMERCIAL",
    "CODICO COMERCIAL": "CODIGO COMERCIAL"
}, inplace=True)

# ─── FILTRA FILAS CON CABECERAS PEGADAS O VACÍAS ───────────────────────────
header_like = set(raw.columns)
header_like.update({"DOC. SUBIDA"})
raw = raw[~raw["COLABORADOR"].str.upper().isin(header_like)]
raw = raw[raw["COLABORADOR"].str.strip() != ""]

# ─── Normaliza campos de texto clave ───────────────────────────────────────
for c in ["PUNTO ATENCION","SERVICIOS","COMUNIDAD","OFERTA PRESENTADA","COLABORADOR"]:
    raw[c] = raw[c].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)


# ─── Anexar CAIDAS de TRAMITACION (cuentan como BAJAS secundarias) ──────────
try:
    wb_tmp  = load_workbook(SRC_XLS, read_only=True)
    sheets  = wb_tmp.sheetnames
    i_tram  = sheets.index("TRAMITACION")
    tram_ss = ["TRAMITACION"] + sheets[i_tram + 1 : i_tram + 3]      # TRAMITACION + 2 sig.

    df_tram = pd.concat(
        [pd.read_excel(SRC_XLS, sheet_name=s, engine="openpyxl") for s in tram_ss],
        ignore_index=True
    )

    # — cabeceras en el mismo formato que «raw» —
    df_tram.columns = [sin_tildes(c).upper().strip() for c in df_tram.columns]

    # normaliza los mismos campos clave
    for col in ["PUNTO ATENCION", "SERVICIOS", "COMUNIDAD",
                "OFERTA PRESENTADA", "COLABORADOR"]:
        if col in df_tram.columns:
            df_tram[col] = (df_tram[col].astype(str)
                                       .str.upper().str.strip()
                                       .str.replace(r"\s+", " ", regex=True))

    # sólo filas con fecha en CAIDAS → son las BAJAS
    df_tram["CAIDAS"] = pd.to_datetime(df_tram["CAIDAS"], errors="coerce")
    df_tram = df_tram[df_tram["CAIDAS"].notna()]

    # alinea columnas que falten / sobren y concatena
    for c in raw.columns.difference(df_tram.columns):
        df_tram[c] = pd.NA
    for c in df_tram.columns.difference(raw.columns):
        raw[c] = pd.NA

    raw = (pd.concat([raw, df_tram], ignore_index=True, sort=False)
             .drop_duplicates())
finally:
    try:
         # evita el bloqueo en Windows
        wb_tmp.close()                     
    except:
        pass

# ─── Mascara “válida” para altas/bajas: descartamos planes BJ/OTROS salvo que tengan servicio ───
mask_plan_invalid     = raw['PLAN'].str.upper().isin(['BJ', 'OTROS'])
mask_serv_ok          = raw['SERVICIOS'].str.upper().str.strip().ne('NO') & raw['SERVICIOS'].notna()
mask_valida_para_alta = ~mask_plan_invalid | mask_serv_ok

from pandas._libs.tslibs.timestamps import Timestamp
raw.loc[~raw["CAIDAS"].apply(lambda x: isinstance(x, Timestamp)), "CAIDAS"] = pd.NaT
raw.loc[~raw["FECHA ALTA"].apply(lambda x: isinstance(x, Timestamp)), "FECHA ALTA"] = pd.NaT
raw["FECHA FIRMA"] = pd.to_datetime(raw["FECHA FIRMA"], errors="coerce")

# -------------- DATES -------------------------------------------------------
if len(sys.argv) >= 3:
    d_ini = pd.to_datetime(sys.argv[1], format="%d-%m-%Y")
    d_fin = pd.to_datetime(sys.argv[2], format="%d-%m-%Y")
else:
    print("⚠️ Sin fechas → pedir.")
    d_ini = ask_date("Desde")
    d_fin = ask_date("Hasta")

HOY = pd.to_datetime(datetime.today().date())       

mask_firma = raw["FECHA FIRMA"].between(d_ini, d_fin, "both")
mask_caida = raw["CAIDAS"].between(d_ini, HOY, "both")
mask_alta_null  = raw["FECHA ALTA"].isna()
mask_caida_null = raw["CAIDAS"].isna()


# ── Recupera texto original de FECHA ALTA ────────────────────────────────
raw["FECHA ALTA ORIGINAL"] = (
    pd.read_excel(SRC_XLS, sheet_name=hoja, engine="openpyxl")["FECHA ALTA"]
)
# Intentamos convertir el valor original a fecha
raw["_FALTA_ORIG_DT"] = pd.to_datetime(
    raw["FECHA ALTA ORIGINAL"], errors="coerce", dayfirst=True
)


ALTAS = raw[mask_firma & raw["CAIDAS"].isna() & mask_valida_para_alta]

mask_incid = (
    mask_firma
    & raw["CAIDAS"].isna()
    & raw["FECHA ALTA"].isna()
    & raw["_FALTA_ORIG_DT"].isna()
    & raw["FECHA ALTA ORIGINAL"].notna()
    & mask_valida_para_alta
)
INCID = raw[mask_incid]

BAJAS = raw[mask_caida & mask_valida_para_alta]

# -------------- TOTAL_GLOBAL ------------------------------------------------
rows = []
raw["CODIGO COMERCIAL"] = raw["CODIGO COMERCIAL"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)


# Códigos comerciales por ubicación

#MIERES
M_CODE = ["YB19010-ANA-3188168", "YB99670-ADRIAN-155292"]
#LENA
L_CODE = ["YB33990-ELI-3189791"]
#PYMES
P_CODE = ["YA8541- GERAR- 3184474"]

def is_mieres(df):
    return df["CODIGO COMERCIAL"].isin(M_CODE)

def is_lena(df):
    return df["CODIGO COMERCIAL"].isin(L_CODE)

def is_pymes(df):
    return df["CODIGO COMERCIAL"].isin(P_CODE) & df["PLAN"].isin(["2,0 TD_3", "3,0 TD"])

def add(tipo, df_a, df_b):
    rows.append({
        "TIPO": tipo,
        "ALTAS": df_a.shape[0],
        "BAJAS": df_b.shape[0],
        "NO_ASTURIAS": df_a[df_a["COMUNIDAD"]!="ASTURIAS"].shape[0],
        "TOTALES": df_a.shape[0] - df_b.shape[0] - df_a[df_a["COMUNIDAD"]!="ASTURIAS"].shape[0],
        "ALTAS_LENA": df_a[is_lena(df_a)].shape[0],
        "BAJAS_LENA": df_b[is_lena(df_b)].shape[0],
        "ALTAS_MIERES": df_a[is_mieres(df_a)].shape[0],
        "BAJAS_MIERES": df_b[is_mieres(df_b)].shape[0],
        "ALTAS_PYMES": df_a[is_pymes(df_a)].shape[0],
        "BAJAS_PYMES": df_b[is_pymes(df_b)].shape[0],

    })


for p in PLANES:
    add(p,
        ALTAS[ALTAS["PLAN"].str.startswith(p,na=False)],
        BAJAS[BAJAS["PLAN"].str.startswith(p,na=False)]
    )
add("Plan Exclusivo 10%",
    ALTAS[contains(ALTAS["OFERTA PRESENTADA"], [OFERTA])],
    BAJAS[contains(BAJAS["OFERTA PRESENTADA"], [OFERTA])]
)
for k,toks in SERVS.items():
    add(k,
        ALTAS[contains(ALTAS["SERVICIOS"], toks)],
        BAJAS[contains(BAJAS["SERVICIOS"], toks)]
    )

add("ALTAS CON INCIDENCIA", INCID, INCID)



total_global = pd.DataFrame(rows)
print(total_global[["TIPO","ALTAS","BAJAS"]])

# -------------- POR_COLAB ---------------------------------------------------
plan_alt = (
    ALTAS[ALTAS["PLAN"].isin(PLANES)]
    .groupby(["COLABORADOR","PLAN"]).size()
    .unstack(fill_value=0).reindex(columns=PLANES, fill_value=0)
)
plan_baj = (
    BAJAS[BAJAS["PLAN"].isin(PLANES)]
    .groupby(["COLABORADOR","PLAN"]).size()
    .unstack(fill_value=0).reindex(columns=PLANES, fill_value=0)
)
plan_alt.columns = [f"PLAN_{c}_ALTA"  for c in plan_alt.columns]
plan_baj.columns = [f"PLAN_{c}_CAIDA" for c in plan_baj.columns]

serv_alt = ALTAS.groupby("COLABORADOR").apply(
    lambda df: pd.Series({
        f"SERVICIO_{k}_ALTA": contains(df["SERVICIOS"], toks).sum()
        for k,toks in SERVS.items()
    })
)
serv_baj = BAJAS.groupby("COLABORADOR").apply(
    lambda df: pd.Series({
        f"SERVICIO_{k}_CAIDA": contains(df["SERVICIOS"], toks).sum()
        for k,toks in SERVS.items()
    })
)

of_alt = (ALTAS[contains(ALTAS["OFERTA PRESENTADA"],[OFERTA])]
          .groupby("COLABORADOR").size().to_frame(f"OFERTA_{OFERTA}_ALTA"))
of_baj = (BAJAS[contains(BAJAS["OFERTA PRESENTADA"],[OFERTA])]
          .groupby("COLABORADOR").size().to_frame(f"OFERTA_{OFERTA}_CAIDA"))

por_colab = (plan_alt.join(plan_baj,how="outer")
                     .join(of_alt,how="outer")
                     .join(of_baj,how="outer")
                     .join(serv_alt,how="outer")
                     .join(serv_baj,how="outer")
                     .fillna(0).astype(int))
por_colab_t = por_colab.T.reset_index()
por_colab_t.columns = ["INDICADOR"] + por_colab_t.columns[1:].tolist()
# ── FILTRA solo indicadores válidos ────────────────────────────────────────
valid = por_colab_t["INDICADOR"].str.match(r"^(PLAN_|OFERTA_|SERVICIO_)")
por_colab_t = por_colab_t[valid].reset_index(drop=True)
# -------------- EXPORT ------------------------------------------------------
out = BASE_DIR / f"Resumen_colaboradores_{datetime.today():%Y-%m-%d}.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    por_colab_t.to_excel(writer, sheet_name="POR_COLABORADOR", index=False)
    total_global.to_excel(writer, sheet_name="TOTAL_GLOBAL",   index=False)
print(f"💾 {out}")

# -------------- FORMAT ------------------------------------------------------
try:
    wb = load_workbook(out)
except PermissionError:
    print(f"❌ No puedo abrir «{out.name}». Asegúrate de que no esté abierto en otro programa.")
    sys.exit(1)
except InvalidFileException:
    print(f"❌ El archivo «{out.name}» no es un libro de Excel válido o está dañado.")
    sys.exit(1)

# Colores y estilos
fills = {
    "head":  PatternFill("solid", fgColor="B7E1CD"),  # verde cabecera
    "alta":  PatternFill("solid", fgColor="C6EFCE"),  # 🟢 verde claro
    "alta_loc": PatternFill("solid", fgColor="C6EFCE"),  # 🟢 verde claro (altas Mieres/Lena/Pymes)
    "baja":  PatternFill("solid", fgColor="FFC7CE"),  # 🔴 rojo claro
    "inci":  PatternFill("solid", fgColor="FFF599"),  # 🟡 amarillo
    "title": PatternFill("solid", fgColor="FBE4D5"),
    "total": PatternFill("solid", fgColor="BDD7EE"),
    "leyenda": PatternFill("solid", fgColor="FFF599"),  # Amarillo específico para leyenda
}
fonts = {"head":Font(bold=True), "title":Font(bold=True, size=12)}
align = Alignment(horizontal="center", vertical="center")
border= Border(*(Side("thin") for _ in range(4)))
per_txt = f"📅 PERÍODO: {d_ini:%d-%m-%Y} → {d_fin:%d-%m-%Y}"

for sh in ["POR_COLABORADOR","TOTAL_GLOBAL"]:
    ws = wb[sh]
    
    # Insertar y formatear título
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    hdr = ws.cell(1, 1, per_txt)
    hdr.fill, hdr.font, hdr.alignment = fills["title"], fonts["title"], align

    # Formatear encabezados
    for c in ws[2]:
        c.fill, c.font, c.alignment, c.border = fills["head"], fonts["head"], align, border

    # Formatear datos (excluyendo la leyenda)
    if sh == "TOTAL_GLOBAL":
        # Primero creamos la leyenda completamente aislada
        BLANK_ROWS = 3
        data_end_row = ws.max_row
        ws.insert_rows(data_end_row + 1, amount=BLANK_ROWS)
        
        # Crear leyenda con formato protegido
        legend_lines = [
            "LEYENDA:",
            "• 🔼 *ALTA*: Firma dentro del período y fecha de CAÍDA vacía.",
            "• 🔽 *BAJA*: Fecha de CAÍDAS dentro del período (independiente de la alta).",
            "• ⚠️ *INCIDENCIA*: Firma dentro del período sin alta válida ni caída.",
            "ℹ️ *RECUERDA*: Las altas con incidencia (RECHAZO, T/A, etc.) se muestran en amarillo y no cuentan como altas ni como bajas."
        ]
        
        # Marcar filas de leyenda para excluirlas del formateo posterior
        legend_start_row = data_end_row + BLANK_ROWS + 1
        legend_rows = set(range(legend_start_row, legend_start_row + len(legend_lines)))
        
        border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
        )
        col_span = ws.max_column


        # Aplicar formato a la leyenda
        for i, text in enumerate(legend_lines):
            row_num = legend_start_row + i
            ws.merge_cells(start_row=row_num, start_column=1, 
                          end_row=row_num, end_column=ws.max_column)
            cell = ws.cell(row=row_num, column=1, value=text)
            cell.fill = fills["leyenda"]
            cell.border = border_thin
            if i == 0:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        
        # Bordes gruesos para la leyenda
        thick = Side(style="medium")
        for col in range(1, ws.max_column + 1):
            top_cell = ws.cell(row=legend_start_row, column=col)
            bot_cell = ws.cell(row=legend_start_row + len(legend_lines) - 1, column=col)
            top_cell.border = Border(top=thick, left=border_thin.left,
                                    right=border_thin.right, bottom=border_thin.bottom)
            bot_cell.border = Border(bottom=thick, left=border_thin.left,
                                     right=border_thin.right, top=border_thin.top)
        
        # Ajustar el rango de formateo para excluir la leyenda
        max_data_row = data_end_row
    else:
        max_data_row = ws.max_row

    # Aplicar formato solo a las filas de datos (excluyendo leyenda)
    for row_cells in ws.iter_rows(min_row=3, max_row=max_data_row):
        first = row_cells[0].value
        if first is None:
            continue
        tipo = first
        for c in row_cells:
            c.alignment, c.border = align, border
            if sh == "POR_COLABORADOR":
                c.fill = fills["alta"] if "_ALTA" in str(row_cells[0].value) else fills["baja"]
            else:
                hdr_txt = ws.cell(row=2, column=c.column).value
                if hdr_txt == "TOTALES":
                    c.fill = fills["total"]
                elif tipo == "ALTAS CON INCIDENCIA":
                    c.fill = fills["inci"]
                elif hdr_txt and hdr_txt.startswith("ALTAS_"):
                    c.fill = fills["alta_loc"]
                elif hdr_txt and hdr_txt.startswith("BAJAS_"):
                    c.fill = fills["baja"]

    auto_width(ws)
    
    if sh == "TOTAL_GLOBAL":
        max_allowed = 25
        if ws.column_dimensions['A'].width > max_allowed:
            ws.column_dimensions['A'].width = max_allowed

wb.save(out)

# -------------- HOJA TRAMITACION ------------------------------------------------------
resp = input("¿Quieres también un informe por colaborador? (S/N): ").strip().upper()
if resp == 'S':
    print("📄 Leyendo la hoja de TRAMITACIÓN …")

    start_blank = ws.max_row + 1
    ws.insert_rows(start_blank, amount=4)
    # ------------------------------------------------- CARGA HOJAS --------------------------------------------------
    def sin_tildes(txt):
        import unicodedata
        return "".join(c for c in unicodedata.normalize("NFKD", txt)
                       if not unicodedata.combining(c))

    df_tram = pd.read_excel(SRC_XLS, sheet_name='TRAMITACION', engine="openpyxl")
    df_tram.columns = [sin_tildes(c).upper().strip() for c in df_tram.columns]

    wb_src       = load_workbook(SRC_XLS, read_only=True)
    names        = wb_src.sheetnames
    i_tram       = names.index("TRAMITACION")
    siguientes   = names[i_tram + 1 : i_tram + 3]

    df_extra = pd.concat(
        [pd.read_excel(SRC_XLS, sheet_name=s, engine="openpyxl") for s in siguientes],
        ignore_index=True
    )
    df_extra.columns = [sin_tildes(c).upper().strip() for c in df_extra.columns]
    for c in ["FECHA FIRMA", "FECHA ALTA", "CAIDAS"]:
        df_extra[c] = pd.to_datetime(df_extra[c], errors="coerce")

    df_tram = pd.concat([df_tram, df_extra], ignore_index=True)
    for c in ["FECHA FIRMA", "FECHA ALTA", "CAIDAS"]:
        df_tram[c] = pd.to_datetime(df_tram[c], errors="coerce")

    # ------------------------------------------------- MÁSCARAS BASE -----------------------------------------------
    mask_plan_no   = df_tram["PLAN"].str.upper().isin(["BJ","OTROS"])
    mask_srv_ok    = df_tram["SERVICIOS"].str.upper().str.strip().ne("NO") & df_tram["SERVICIOS"].notna()
    mask_valida    = ~mask_plan_no | mask_srv_ok

    wb     = load_workbook(out)
    hoy    = pd.to_datetime(datetime.today().date())

    fills  = {
        "alta": PatternFill("solid", fgColor="D5F5D3"),
        "baja": PatternFill("solid", fgColor="FFC7CE"),
        "inci": PatternFill("solid", fgColor="FFF2CC"),
        "sec" : PatternFill("solid", fgColor="C9DAF8"),
    }
    hdr_fill   = PatternFill("solid", fgColor="E3E4FA")    # lavanda pálido
    hdr_font   = Font(bold=True, size=12)
    border_thin = Border(*(Side("thin") for _ in range(4)))

    # ---------- columnas definitivas ---------------------------------------------------
    cols = [
        "INDICE","COLABORADOR","NOMBRE DEL CLIENTE","DNI/CIF",
        "PLAN","POTENCIA","OFERTA PRESENTADA","SERVICIOS",
        "FECHA FIRMA","FECHA ALTA","OBSERV.","CAIDAS","CHECK ALTAS",
        "VIENE GRACIAS A :","OTROS"
    ]

    for col in df_tram["COLABORADOR"].dropna().unique():
        nombre = re.sub(r"[\\/?*\[\]]","_", str(col).strip()[:31])

        m_col  = df_tram["COLABORADOR"].astype(str).str.strip().str.upper() == str(col).strip().upper()
        m_alta = df_tram["FECHA FIRMA"].between(d_ini, d_fin, "both") & df_tram["CAIDAS"].isna() & mask_valida
        m_baja = df_tram["CAIDAS"].between(d_ini, d_fin, "both") & mask_valida
        m_inci = m_alta & df_tram["FECHA ALTA"].isna()
        m_sec  = (df_tram["FECHA FIRMA"] < d_ini) & df_tram["CAIDAS"].between(d_ini, hoy, "both") & mask_valida

        sel    = m_col & (m_alta | m_baja | m_inci | m_sec)
        df_fil = df_tram.loc[sel].copy()
        df_fil.insert(0, "INDICE", range(1, len(df_fil) + 1))   # columna numerada
        df_fil = df_fil[cols]

        ws = wb.create_sheet(title=nombre)

        # -------------- CABECERA -------------------------------------------------------
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       text_rotation=90,
                                       wrap_text=True)
        ws.row_dimensions[1].height = 80   # alto suficiente

        # -------------- DATOS ----------------------------------------------------------
        for ri, (idx_real, row) in enumerate(df_fil.iterrows(), start=2):
            if   m_sec.loc[idx_real]:  fill = fills["sec"]
            elif m_inci.loc[idx_real]: fill = fills["inci"]
            elif m_baja.loc[idx_real]: fill = fills["baja"]
            else:                       fill = fills["alta"]

            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill      = fill
                c.alignment = Alignment(horizontal="center", vertical="center")

        auto_width(ws)

  # — Crear o limpiar la hoja de leyenda —
if "LEYENDA" in wb.sheetnames:
    ws_ley = wb["LEYENDA"]
    # si ya existía, borra todo su contenido:
    for row in ws_ley["A1:D4"]:
        for cell in row:
            cell.value = None
else:
    ws_ley = wb.create_sheet(title="LEYENDA")

# Define la leyenda global
leyenda = [
    ('🟩 ALTA',       "Firma entre fecha de inicio y fecha fin, sin caída",      'alta'),
    ('🟥 BAJA',       "Caída entre fecha de inicio y fecha fin",                 'baja'),
    ('🟨 INCIDENCIA', "Firma en rango, sin alta ni caída",                      'inci'),
    ('🟦 SECUNDARIO', "Firma < fecha de inicio y caída entre inicio y hoy",    'sec'),
]

# Estilos base
fills_ley = {
    'alta': PatternFill("solid", fgColor="D5F5D3"),
    'baja': PatternFill("solid", fgColor="FFC7CE"),
    'inci': PatternFill("solid", fgColor="FFF2CC"),
    'sec' : PatternFill("solid", fgColor="C9DAF8"),
}
border_top = Border(top=Side(style="medium"))
border_bot = Border(bottom=Side(style="medium"))
font_ital  = Font(italic=True)

# Escribe las 4 filas de leyenda en A1:D4
for i, (lbl, desc, key) in enumerate(leyenda, start=1):
    ws_ley.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
    cell = ws_ley.cell(row=i, column=1, value=f"{lbl}: {desc}")
    cell.fill      = fills_ley[key]
    cell.font      = font_ital
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_ley.row_dimensions[i].height = 20
    # bordes gruesos en primera y última fila
    if i == 1:
        cell.border = border_top
    elif i == len(leyenda):
        cell.border = border_bot

# Ajusta el ancho de columnas A–D para que el texto quepa
for col in range(1, 5):
    ws_ley.column_dimensions[get_column_letter(col)].width = 50


try:
    wb.save(out)
except PermissionError:
    print(f"❌ No puedo guardar «{out.name}». Cierra el archivo si está abierto y vuelve a intentarlo.")
    sys.exit(1)

import os; os.startfile(out) 

print("✅ Fin.")
